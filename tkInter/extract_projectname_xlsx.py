
import ifcopenshell
import openpyxl
from tkinter import filedialog
import tkinter as tk

def extract_project_name(ifc_file_path):
    print(ifc_file_path)
    ifc_file = ifcopenshell.open(ifc_file_path)

    # Iterate over each schema in the IFC file
    for item in ifc_file.by_type("IfcProject"):
            return item.Name

    return None

def write_to_excel(project_name):
    excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    if excel_path:
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="Projektname")
        ws.cell(row=2, column=1, value=project_name)

        wb.save(excel_path)
        print(f"Projektnamen in Excel geschrieben: {excel_path}")

def choose_ifc_file_and_process():
    ifc_file_path = filedialog.askopenfilename(filetypes=[("IFC files", "*.ifc")])

    if ifc_file_path:
        project_name = extract_project_name(ifc_file_path)
        if project_name:
            write_to_excel(project_name)
        else:
            print("Projektname konnte nicht gefunden werden.")
    else:
        print("IFC-Datei nicht ausgewählt.")

root = tk.Tk()
root.title("IFC Projektname in Excel schreiben")

btn_choose_ifc_and_process = tk.Button(root, text="IFC-Datei auswählen und verarbeiten", command=choose_ifc_file_and_process)
btn_choose_ifc_and_process.pack(pady=50)

root.mainloop()
